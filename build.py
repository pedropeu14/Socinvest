#!/usr/bin/env python3
"""Gera data.json a partir do export Bloomberg (OB_OS.xlsx).

Estrutura esperada em cada aba:
  Coluna A: Date | Coluna B: MOV_AVG_200D | Coluna D: Date | Coluna E: PX_LAST
  O ticker é lido da primeira célula de texto da coluna A (ex.: "SPY US EQUITY").

Robusto a: datas armazenadas como texto (vários formatos), números como texto
(ponto ou vírgula decimal). O formato de data é detectado por coluna, exigindo
que todas as células parseiem e que a série fique em ordem cronológica.

Uso:  python build.py [caminho_do_xlsx]   (default: OB_OS.xlsx)
"""
import sys, json, statistics
from datetime import datetime, date
import openpyxl

SKIP_SHEETS = {"Charts"}
DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"]

def to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("'", "")
        if s.count(",") == 1 and s.count(".") == 0:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None

def parse_date_column(raw):
    """raw: lista de valores brutos. Retorna lista de date|None do mesmo tamanho."""
    out = [v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None) for v in raw]
    texts = [(i, v.strip()) for i, v in enumerate(raw) if isinstance(v, str) and v.strip() and "date" not in v.lower()]
    texts = [(i, s) for i, s in texts if any(c.isdigit() for c in s)]
    if not texts:
        return out
    best = None
    for fmt in DATE_FORMATS:
        parsed = []
        ok = True
        for i, s in texts:
            try:
                parsed.append((i, datetime.strptime(s[:10], fmt).date()))
            except ValueError:
                ok = False
                break
        if not ok:
            continue
        merged = list(out)
        for i, d in parsed:
            merged[i] = d
        seq = [d for d in merged if d]
        if seq == sorted(seq):  # série cronológica => formato correto
            best = parsed
            break
        if best is None:
            best = parsed  # fallback: primeiro formato que parseia tudo
    if best:
        for i, d in best:
            out[i] = d
    return out

def main(path="OB_OS.xlsx"):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        sh = wb[name]
        ticker = None
        colA, colB, colD, colE = [], [], [], []
        for r in sh.iter_rows(min_row=1, max_col=5, values_only=True):
            a, b, _, d, e = (list(r) + [None] * 5)[:5]
            if ticker is None and isinstance(a, str) and "Date" not in a and a.strip() and not any(c.isdigit() for c in a):
                ticker = a.strip()
            colA.append(a); colB.append(b); colD.append(d); colE.append(e)
        datesA = parse_date_column(colA)
        datesD = parse_date_column(colD)
        ma = {dt: to_number(v) for dt, v in zip(datesA, colB) if dt and to_number(v) is not None}
        px = {dt: to_number(v) for dt, v in zip(datesD, colE) if dt and to_number(v) is not None}
        common = sorted(set(ma) & set(px))
        series = [(dt, px[dt] / ma[dt]) for dt in common if ma[dt] != 0]
        if not series:
            print(f"AVISO: aba '{name}' sem dados válidos — ignorada", file=sys.stderr)
            continue
        vals = [v for _, v in series]
        out[name] = {
            "ticker": ticker or name,
            "dates": [dt.isoformat() for dt, _ in series],
            "ratio": [round(v, 4) for v in vals],
            "mean": round(statistics.mean(vals), 4),
            "sd": round(statistics.pstdev(vals), 4),
            "last": round(vals[-1], 4),
            "lastDate": series[-1][0].isoformat(),
            "n": len(vals),
        }
        print(f"{name:12s} {ticker or '?':22s} {len(vals):>5d} obs  até {series[-1][0]}")
    with open("data.json", "w") as f:
        json.dump(out, f)
    print(f"\ndata.json gerado: {len(out)} ativos")



# ---------- P/E (formato wide: p_e.xlsx) ----------
PE_FILE = "p_e.xlsx"
MIN_OBS_PE = 30

def build_pe(path=PE_FILE):
    import os
    if not os.path.exists(path):
        print(f"{path} não encontrado — pe.json não gerado (aba Valuation ficará oculta)")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = [list(r) for r in sh.iter_rows(values_only=True)]
    # localizar linha de tickers: a que tem mais células contendo Equity/Index/Curncy
    hdr_i, hdr_score = None, 0
    for i, r in enumerate(rows[:15]):
        score = sum(1 for v in r if isinstance(v, str) and any(k in v.lower() for k in ("equity", "index", "curncy", "comdty")))
        if score > hdr_score:
            hdr_i, hdr_score = i, score
    if hdr_i is None:
        print("AVISO: linha de tickers não encontrada no arquivo de P/E", file=sys.stderr)
        return
    headers = rows[hdr_i]
    datesA = parse_date_column([r[0] if r else None for r in rows])
    out = {}
    for j, t in enumerate(headers):
        if not isinstance(t, str) or not t.strip() or j == 0:
            continue
        t = t.strip()
        series = []
        for i, r in enumerate(rows):
            if not datesA[i] or j >= len(r):
                continue
            v = to_number(r[j])
            if v is not None and v > 0:
                series.append((datesA[i], v))
        if len(series) < MIN_OBS_PE:
            if series or any(isinstance(r[j] if j < len(r) else None, str) for r in rows[hdr_i+1:]):
                print(f"P/E: '{t}' ignorado ({len(series)} obs — insuficiente)", file=sys.stderr)
            continue
        series.sort()
        vals = [v for _, v in series]
        # Estatística robusta (trailing P/E tem picos extremos quando o lucro colapsa):
        # mediana no lugar da média; (P84-P16)/2 como equivalente robusto de 1 sigma.
        sv = sorted(vals)
        def pct(p):
            k = (len(sv) - 1) * p
            f = int(k)
            return sv[f] + (sv[min(f + 1, len(sv) - 1)] - sv[f]) * (k - f)
        med, p16, p84 = pct(0.5), pct(0.16), pct(0.84)
        out[t] = {
            "ticker": t.upper(),
            "dates": [d.isoformat() for d, _ in series],
            "ratio": [round(v, 3) for v in vals],
            "mean": round(med, 3),
            "sd": round((p84 - p16) / 2, 3),
            "last": round(vals[-1], 3),
            "lastDate": series[-1][0].isoformat(),
            "n": len(vals),
            "robust": True,
        }
        print(f"P/E  {t:22s} {len(vals):>5d} obs  {series[0][0]} → {series[-1][0]}")
    with open("pe.json", "w") as f:
        json.dump(out, f)
    print(f"pe.json gerado: {len(out)} ativos")



# ---------- Catálogo de ETFs (Catalogo_de_ETFs.xlsx) ----------
CAT_FILE = "Catalogo_de_ETFs.xlsx"

def build_catalog(path=CAT_FILE):
    import os, re
    if not os.path.exists(path):
        print(f"{path} não encontrado — catalogo.json não gerado (aba Catálogo ficará oculta)")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"etfs": [], "overlap": {"tickers": [], "rows": []}}

    # --- Índice ---
    if "Indice" in wb.sheetnames:
        rows = [list(r) for r in wb["Indice"].iter_rows(values_only=True)]
        hdr = next((i for i, r in enumerate(rows) if r and r[0] == "Ticker"), None)
        if hdr is not None:
            for r in rows[hdr + 1:]:
                if not r or not r[0]:
                    continue
                out["etfs"].append({
                    "ticker": str(r[0]).strip(),
                    "nome": str(r[1] or "").strip(),
                    "emissor": str(r[2] or "").strip(),
                    "categoria": str(r[3] or "").strip(),
                    "domicilio": str(r[4] or "").strip(),
                    "moeda": str(r[5] or "").strip(),
                    "ter": to_number(r[6]),
                    "descricao": "", "holdings": [], "fonte": "",
                })

    # --- Abas por ETF ---
    by_key = {" ".join(e["ticker"].split()[:2]).upper(): e for e in out["etfs"]}
    for name in wb.sheetnames:
        if name in ("Indice", "Overlap"):
            continue
        key = name.replace("_", " ").upper()
        e = by_key.get(key)
        if e is None:
            continue
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
        mode = None
        for r in rows:
            a = r[0] if r else None
            if isinstance(a, str):
                s = a.strip()
                if s.upper().startswith("O QUE E"):
                    mode = "desc"; continue
                if s.upper().startswith("HOLDINGS"):
                    mode = None; continue
                if s == "#":
                    mode = "hold"; continue
                if s.startswith("Fonte:"):
                    e["fonte"] = s; mode = None; continue
            if mode == "desc" and isinstance(a, str) and a.strip():
                e["descricao"] = (e["descricao"] + " " + a.strip()).strip()
            elif mode == "hold" and isinstance(a, (int, float)):
                e["holdings"].append({
                    "ativo": str(r[1] or "").strip(),
                    "tic": str(r[2] or "").strip(),
                    "peso": to_number(r[3]),
                })
        print(f"Catálogo: {e['ticker']:18s} {len(e['holdings']):>3d} holdings")

    # --- Overlap ---
    if "Overlap" in wb.sheetnames:
        rows = [list(r) for r in wb["Overlap"].iter_rows(values_only=True)]
        hdr = next((i for i, r in enumerate(rows) if r and r[0] == "Ticker"), None)
        if hdr is not None:
            cols = [str(c).strip() for c in rows[hdr][3:] if c]
            out["overlap"]["tickers"] = cols
            for r in rows[hdr + 1:]:
                if not r or not r[0] or str(r[0]).startswith("Nota"):
                    continue
                pesos = [to_number(r[3 + j]) if 3 + j < len(r) else None for j in range(len(cols))]
                if not any(p is not None for p in pesos):
                    continue
                out["overlap"]["rows"].append({
                    "tic": str(r[0]).strip(), "ativo": str(r[1] or "").strip(),
                    "n": int(to_number(r[2]) or 0), "pesos": pesos,
                })
    with open("catalogo.json", "w") as f:
        json.dump(out, f)
    print(f"catalogo.json gerado: {len(out['etfs'])} ETFs, {len(out['overlap']['rows'])} linhas de overlap")

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "OB_OS.xlsx")
    build_pe(sys.argv[2] if len(sys.argv) > 2 else PE_FILE)
    build_catalog(sys.argv[3] if len(sys.argv) > 3 else CAT_FILE)

