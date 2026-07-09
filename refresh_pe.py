#!/usr/bin/env python3
"""Atualiza o p_e.xlsx (abas Trailing + Forward) via Desktop API Bloomberg — por MERGE.

Roda APENAS na máquina com o Terminal Bloomberg aberto e logado (Desktop API
em localhost:8194 — o mesmo entitlement do add-in do Excel). NÃO roda em CI.
Se o Terminal estiver fechado/deslogado, sai com erro sem tocar em nada — o
site continua andando pela extensão de preço diária do CI.

Filosofia: MERGE, nunca reconstrução. A planilha atual é a memória (inclusive
histórico que a API não entrega mais — a profundidade do consenso via DAPI é
menor que a do export original do add-in). O refresh busca as últimas semanas
e sobrepõe/acrescenta; é impossível perder colunas ou história por construção.
Cada aba mantém seus PRÓPRIOS tickers (a Trailing usa 2330 TT/VALE3 BZ/BRK/A,
onde a Bloomberg tem PE_RATIO; a Forward usa as linhas US).

Campos (validados contra o export do add-in, valores idênticos):
  Trailing PE -> PE_RATIO          (P/E dos últimos 12 meses)
  Forward PE  -> BEST_PE_RATIO     (P/E do consenso 12 meses à frente)

Fluxo completo (refresh_pe.cmd faz os três passos):
    python refresh_pe.py
    git add p_e.xlsx && git commit -m "data: P/E refresh" && git push
    # o push dispara o CI -> rebuild pe.json/pe_fwd.json -> re-ancora a extensão

Requer: pip install openpyxl blpapi
  (blpapi: pip install blpapi --index-url
   https://blpapi.bloomberg.com/repository/releases/python/simple/)
"""
import sys
from datetime import date, datetime, timedelta

import blpapi
import openpyxl

PE_FILE = "p_e.xlsx"
# (nome da aba no xlsx, campo Bloomberg)
SHEETS = [("Forward PE", "BEST_PE_RATIO"), ("Trailing PE", "PE_RATIO")]
LOOKBACK_DAYS = 180     # rebusca ~6 meses: cobre revisões retroativas da Bloomberg


def norm(t: str) -> str:
    return " ".join(t.upper().split()[:2])


def read_sheet(wb, sheet: str) -> tuple[list[str], dict[str, dict[date, float]]]:
    """Tickers (na ordem das colunas) e séries {ticker: {data: valor}} da aba."""
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    hdr_i, score = None, 0
    for i, r in enumerate(rows[:15]):
        n = sum(1 for v in r if isinstance(v, str)
                and any(k in v.lower() for k in ("equity", "index", "curncy", "comdty")))
        if n > score:
            hdr_i, score = i, n
    if hdr_i is None:
        sys.exit(f"ERRO: linha de tickers não encontrada na aba '{sheet}'")
    tickers, cols = [], []
    for j, v in enumerate(rows[hdr_i]):
        if isinstance(v, str) and v.strip() and "date" not in v.lower():
            tickers.append(v.strip())
            cols.append(j)
    series: dict[str, dict[date, float]] = {t: {} for t in tickers}
    for r in rows[hdr_i + 1:]:
        d = r[0] if r else None
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            continue
        for t, j in zip(tickers, cols):
            if j < len(r) and isinstance(r[j], (int, float)):
                series[t][d] = float(r[j])
    return tickers, series


def fetch_recent(tickers: list[str], field: str, start: date) -> dict[str, dict[date, float]]:
    """Semanas recentes de um campo para uma lista de tickers via DAPI."""
    opts = blpapi.SessionOptions()
    opts.setServerHost("localhost")
    opts.setServerPort(8194)
    session = blpapi.Session(opts)
    if not session.start():
        sys.exit("ERRO: Desktop API não respondeu — o Terminal está aberto e logado?")
    try:
        if not session.openService("//blp/refdata"):
            sys.exit("ERRO: //blp/refdata indisponível")
        svc = session.getService("//blp/refdata")
        req = svc.createRequest("HistoricalDataRequest")
        for t in tickers:
            req.getElement("securities").appendValue(t)
        req.getElement("fields").appendValue(field)
        req.set("startDate", start.strftime("%Y%m%d"))
        req.set("endDate", date.today().strftime("%Y%m%d"))
        req.set("periodicitySelection", "WEEKLY")
        session.sendRequest(req)

        out: dict[str, dict[date, float]] = {}
        errors: list[str] = []
        while True:
            ev = session.nextEvent(60000)
            if ev.eventType() == blpapi.Event.TIMEOUT:
                sys.exit("ERRO: timeout esperando resposta da API")
            for msg in ev:
                if not msg.hasElement("securityData"):
                    continue
                sd = msg.getElement("securityData")
                tick = sd.getElementAsString("security")
                if sd.hasElement("securityError"):
                    errors.append(f"{tick}: "
                                  + sd.getElement("securityError").getElementAsString("message"))
                    continue
                s = out.setdefault(tick, {})
                fd = sd.getElement("fieldData")
                for i in range(fd.numValues()):
                    row = fd.getValueAsElement(i)
                    if row.hasElement(field):
                        d = row.getElementAsDatetime("date")
                        d = d.date() if isinstance(d, datetime) else d
                        s[d] = row.getElementAsFloat(field)
            if ev.eventType() == blpapi.Event.RESPONSE:
                break
        if errors:
            sys.exit("ERRO (p_e.xlsx preservado):\n  " + "\n  ".join(errors))
        return out
    finally:
        session.stop()


def main() -> int:
    wb_in = openpyxl.load_workbook(PE_FILE, read_only=True, data_only=True)
    merged: dict[str, tuple[list[str], dict[str, dict[date, float]]]] = {}
    start = date.today() - timedelta(days=LOOKBACK_DAYS)
    for sheet, field in SHEETS:
        if sheet not in wb_in.sheetnames:
            print(f"AVISO: aba '{sheet}' não existe no arquivo — pulada")
            continue
        tickers, old = read_sheet(wb_in, sheet)
        fresh = fetch_recent(tickers, field, start)
        fresh_by_key = {norm(t): s for t, s in fresh.items()}
        added = 0
        for t in tickers:
            new = fresh_by_key.get(norm(t), {})
            before = len(old[t])
            old[t].update(new)          # DAPI ganha onde há sobreposição
            added += len(old[t]) - before
        merged[sheet] = (tickers, old)
        last = max((max(s) for s in old.values() if s), default="?")
        print(f"{sheet}: {len(tickers)} tickers, +{added} observações novas, até {last}")
    if not merged:
        sys.exit("ERRO: nenhuma aba reconhecida no p_e.xlsx")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    for sheet, _ in SHEETS:
        if sheet not in merged:
            continue
        tickers, series = merged[sheet]
        all_dates = sorted({d for s in series.values() for d in s})
        sh = wb_out.create_sheet(sheet)
        sh.cell(row=1, column=1, value="Date")
        for j, t in enumerate(tickers, start=2):
            sh.cell(row=1, column=j, value=t)
        for i, d in enumerate(all_dates, start=2):
            sh.cell(row=i, column=1, value=datetime(d.year, d.month, d.day))
            for j, t in enumerate(tickers, start=2):
                v = series[t].get(d)
                if v is not None:
                    sh.cell(row=i, column=j, value=round(v, 4))
    wb_out.save(PE_FILE)
    print(f"{PE_FILE} atualizado (merge — histórico preservado).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
